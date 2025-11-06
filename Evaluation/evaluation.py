
"""

Strict evaluation script
Maatch predicted ingredients to ground truth ingredients using improved fuzzy matching,
use both visible and hidden ingredients for matching,
excludes certain ground truth entries based on keywords to avoid ambiguity
"deprecated", "Mixed green", "plate only", "salad"
and generate a detailed Excel report comparing quantities
"""
import json
import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re

# ====== File Paths ======
GROUND_TRUTH_PATH = "Nutrition5k/metadata/dish_metadata_cafe1.json"
PRED_PATH = "output/Nutrition5k_Gemini-2.5-pro_pydantic_food_dataset_analysis.json"  #or "output/Nutrition5k_700_800_Gemini-2.5-pro_pydantic_food_dataset_analysis.json"
OUTPUT_XLSX = r"Evaluation/strict_ingredient_comparison_groundtruth_based.xlsx"

# ====== Fuzzy Matching Threshold ======
FUZZY_THRESHOLD = 80

# ====== Read JSON Files ======
with open(GROUND_TRUTH_PATH, "r", encoding="utf-8") as f:
    gt_data = json.load(f)
with open(PRED_PATH, "r", encoding="utf-8") as f:
    pred_data = json.load(f)

# ====== Build Prediction Lookup Table ======
pred_dict = {p["image_filename"]: p for p in pred_data}

# ====== Filter Function ======
def should_skip_ground_truth(gt_entry) -> bool:
    """Check if this ground truth entry should be skipped"""
    skip_keywords = ["deprecated", "Mixed green", "plate only", "salad"]
    
    # Check if ingredients contain any skip keywords
    ingredients = gt_entry.get("ingredients", [])
    for ingredient in ingredients:
        ingredient_name = ingredient.get("name", "").lower()
        for keyword in skip_keywords:
            if keyword.lower() in ingredient_name:
                return True
    
    return False

# ====== Enhanced Normalization Function ======
def normalize_name(name: str) -> str:
    """Enhanced text normalization"""
    if not name:
        return ""
    name = name.lower().strip()
    name = re.sub(r"[^a-z\s]", "", name)  # Remove symbols
    # Handle common plurals
    plural_map = {
        "berries": "berry", "grapes": "grape", "olives": "olive",
        "tomatoes": "tomato", "potatoes": "potato"
    }
    for plural, singular in plural_map.items():
        name = name.replace(plural, singular)
    name = " ".join(name.split())  # Normalize whitespace
    return name

# ====== Fuzzy Matching Function ======
def fuzzy_match(gt_name, pred_names):
    """Return the most similar predicted ingredient name and score"""
    if not pred_names:
        return None, 0
    match, score, _ = process.extractOne(gt_name, pred_names, scorer=fuzz.token_sort_ratio)
    return match, score

# ====== Optimized Matching Logic ======
def find_matching_predictions(gt_name, pred_visible_names, pred_visible_list, pred_hidden_names, pred_hidden_list):
    """Optimized matching logic"""
    normalized_gt = normalize_name(gt_name)
    
    matching_items = []
    match_type = "❌ No Match"
    score = 0
    
    # 1️⃣ First try fuzzy matching with visible ingredients
    match_name, score = fuzzy_match(normalized_gt, pred_visible_names)
    
    if match_name and score >= FUZZY_THRESHOLD:
        # High score match - directly adopt
        matching_items = [p for p in pred_visible_list if normalize_name(p["name"]) == match_name]
        match_type = "✅ High"
        return matching_items, score, match_type
    
    elif match_name and score > 0:
        # Low score match - check containment relationship
        for pred_item in pred_visible_list:
            pred_norm = normalize_name(pred_item["name"])
            if (normalized_gt in pred_norm or pred_norm in normalized_gt):
                if pred_item not in matching_items:
                    matching_items.append(pred_item)
        
        if matching_items:
            match_type = "🟡 Contained"
            return matching_items, score, match_type
    
    # 2️⃣ If no match in visible, try hidden ingredients
    if not matching_items:
        match_name, hidden_score = fuzzy_match(normalized_gt, pred_hidden_names)
        
        if match_name and hidden_score >= FUZZY_THRESHOLD:
            # High score match in hidden
            matching_items = [p for p in pred_hidden_list if normalize_name(p["name"]) == match_name]
            match_type = "✅ High (Hidden)"
            return matching_items, hidden_score, match_type
        
        elif match_name and hidden_score > 0:
            # Low score match in hidden - check containment
            for pred_item in pred_hidden_list:
                pred_norm = normalize_name(pred_item["name"])
                if (normalized_gt in pred_norm or pred_norm in normalized_gt):
                    if pred_item not in matching_items:
                        matching_items.append(pred_item)
            
            if matching_items:
                match_type = "🟡 Contained (Hidden)"
                return matching_items, hidden_score, match_type
    
    return matching_items, score, match_type

# ====== Main Loop ======
rows = []
skipped_count = 0
processed_count = 0

for filename, pred_entry in pred_dict.items():
    # Find corresponding ground truth
    gt_entry = next((g for g in gt_data if g["image_filename"] == filename), None)
    if not gt_entry:
        continue
    
    # Check if ground truth contains skip keywords
    if should_skip_ground_truth(gt_entry):
        skipped_count += 1
        continue
    
    processed_count += 1
    visible_pred = pred_entry.get("visible_ingredients", [])
    hidden_pred = pred_entry.get("hidden_ingredients", [])
    gt_ings = gt_entry.get("ingredients", [])
    
    pred_visible_names = [normalize_name(p["name"]) for p in visible_pred]
    pred_hidden_names = [normalize_name(p["name"]) for p in hidden_pred]

    for gt_item in gt_ings:
        gt_name = gt_item["name"]
        gt_qty = float(gt_item.get("quantity", 0))

        # Use optimized matching logic (now includes hidden ingredients)
        matching_items, score, confidence = find_matching_predictions(
            gt_name, pred_visible_names, visible_pred, pred_hidden_names, hidden_pred
        )

        if confidence == "❌ No Match":
            # No Match case: keep only basic info, leave numeric fields empty
            rows.append({
                "image_filename": filename,
                "gt_name": gt_name,
                "pred_name": None,
                "similarity_score": round(score, 1),
                "gt_qty": None,  # Leave empty
                "pred_qty": None,  # Leave empty
                "diff": None,  # Leave empty
                "abs_diff": None,  # Leave empty
                "pct_diff": None,  # Leave empty
                "confidence": confidence
            })
            continue

        # Handle multiple matches
        total_pred_qty = sum(float(p.get("quantity", 0)) for p in matching_items)
        pred_names_combined = " + ".join([p["name"] for p in matching_items])
        
        # Calculate differences
        diff = total_pred_qty - gt_qty
        abs_diff = abs(diff)
        pct_diff = (diff / gt_qty * 100) if gt_qty != 0 else None

        rows.append({
            "image_filename": filename,
            "gt_name": gt_name,
            "pred_name": pred_names_combined,
            "similarity_score": round(score, 1),
            "gt_qty": gt_qty,
            "pred_qty": round(total_pred_qty, 2),
            "diff": round(diff, 2),
            "abs_diff": round(abs_diff, 2),
            "pct_diff": round(pct_diff, 1) if pct_diff is not None else None,
            "confidence": confidence
        })

# ====== Export to Excel ======
df = pd.DataFrame(rows)
df.to_excel(OUTPUT_XLSX, index=False)

# ====== Excel Formatting ======
wb = load_workbook(OUTPUT_XLSX)
ws = wb.active

# Auto-adjust column widths
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# Merge cells with same filename
col_idx = 1
start_row = 2
for i in range(2, ws.max_row + 2):
    if i == ws.max_row + 1 or ws.cell(i, col_idx).value != ws.cell(start_row, col_idx).value:
        if i - start_row > 1:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=i-1, end_column=col_idx)
            ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
        start_row = i

wb.save(OUTPUT_XLSX)
wb.close()

print(f"✅ Completed! Processed {processed_count} images, skipped {skipped_count} images containing filter keywords")
print(f"📊 Generated {len(df)} rows, saved to: {OUTPUT_XLSX}")


##########################################################################################################

"""Raw evaluation script"""



# import json
# import pandas as pd
# from rapidfuzz import process, fuzz
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Alignment

# # ====== File Paths ======
# GROUND_TRUTH_PATH = "Nutrition5k/metadata/dish_metadata_cafe1.json"
# PRED_PATH = "output/Nutrition5k_Gemini-2.5-pro_pydantic_food_dataset_analysis.json"  #or "output/Nutrition5k_700_800_Gemini-2.5-pro_pydantic_food_dataset_analysis.json"
# OUTPUT_XLSX = r"Evaluation/strict_ingredient_comparison_groundtruth_based.xlsx"

# # ====== Fuzzy Matching Threshold ======
# FUZZY_THRESHOLD = 80

# # ====== Read JSON Files ======
# with open(GROUND_TRUTH_PATH, "r", encoding="utf-8") as f:
#     gt_data = json.load(f)
# with open(PRED_PATH, "r", encoding="utf-8") as f:
#     pred_data = json.load(f)

# # ====== Build Prediction Lookup Table ======
# pred_dict = {p["image_filename"]: p for p in pred_data}

# # ====== Improved Fuzzy Matching Functions ======
# def find_best_match_for_gt(gt_name, pred_ingredients):
#     """Find the best matching predicted ingredient for a single ground truth ingredient"""
#     if not pred_ingredients:
#         return None, 0, None
    
#     pred_names = [p["name"].lower().strip() for p in pred_ingredients]
    
#     # First try exact matching
#     for i, pred_name in enumerate(pred_names):
#         if gt_name == pred_name:
#             return pred_name, 100, pred_ingredients[i]
    
#     # Try containment matching
#     for i, pred_name in enumerate(pred_names):
#         if gt_name in pred_name or pred_name in gt_name:
#             return pred_name, 95, pred_ingredients[i]
    
#     # Fuzzy matching
#     result = process.extractOne(
#         gt_name, 
#         pred_names, 
#         scorer=fuzz.token_sort_ratio,
#         score_cutoff=60
#     )
    
#     if result:
#         match, score, idx = result
#         if score >= 60:
#             return match, score, pred_ingredients[idx]
    
#     return None, 0, None

# def find_all_matches(gt_ingredients, pred_ingredients):
#     """Find best matches for all ground truth ingredients, avoiding conflicts"""
#     matches = []
#     used_pred_indices = set()
    
#     # First round: match high confidence ones first (exact and containment matches)
#     for gt_idx, gt_item in enumerate(gt_ingredients):
#         gt_name = gt_item["name"].lower().strip()
        
#         # Get all available predicted ingredients
#         available_pred = [p for i, p in enumerate(pred_ingredients) if i not in used_pred_indices]
        
#         # Find best match
#         match_name, score, pred_item = find_best_match_for_gt(gt_name, available_pred)
        
#         if pred_item:
#             pred_idx = pred_ingredients.index(pred_item)
#             # Only occupy immediately for high confidence matches
#             if score >= 85:
#                 used_pred_indices.add(pred_idx)
#                 matches.append((gt_idx, pred_idx, score))
    
#     # Second round: match all remaining ingredients
#     for gt_idx, gt_item in enumerate(gt_ingredients):
#         # If this ground truth ingredient hasn't been matched yet
#         if not any(m[0] == gt_idx for m in matches):
#             gt_name = gt_item["name"].lower().strip()
            
#             # Get all available predicted ingredients
#             available_pred = [p for i, p in enumerate(pred_ingredients) if i not in used_pred_indices]
            
#             # Find best match
#             match_name, score, pred_item = find_best_match_for_gt(gt_name, available_pred)
            
#             if pred_item:
#                 pred_idx = pred_ingredients.index(pred_item)
#                 used_pred_indices.add(pred_idx)
#                 matches.append((gt_idx, pred_idx, score))
    
#     return matches

# # ====== Main Loop ======
# rows = []

# for filename, pred_entry in pred_dict.items():
#     visible_pred = pred_entry.get("visible_ingredients", [])
#     hidden_pred = pred_entry.get("hidden_ingredients", [])
    
#     # Combine all predicted ingredients but keep source information
#     all_pred = []
#     for ing in visible_pred:
#         ing['source'] = 'visible'
#         all_pred.append(ing)
#     for ing in hidden_pred:
#         ing['source'] = 'hidden' 
#         all_pred.append(ing)

#     # Find ground truth
#     gt_entry = next((g for g in gt_data if g["image_filename"] == filename), None)
#     if not gt_entry:
#         continue

#     gt_ings = gt_entry.get("ingredients", [])
    
#     # Find all best matches
#     matches = find_all_matches(gt_ings, all_pred)
    
#     # Process matching results
#     matched_gt_indices = set()
#     matched_pred_indices = set()
    
#     for gt_idx, pred_idx, score in matches:
#         matched_gt_indices.add(gt_idx)
#         matched_pred_indices.add(pred_idx)
        
#         gt_item = gt_ings[gt_idx]
#         pred_item = all_pred[pred_idx]
        
#         gt_name = gt_item["name"]
#         gt_qty = float(gt_item.get("quantity", 0))
#         pred_qty = float(pred_item.get("quantity", 0))
        
#         # Calculate differences
#         diff = pred_qty - gt_qty
#         abs_diff = abs(diff)
#         pct_diff = (diff / gt_qty * 100) if gt_qty != 0 else None
        
#         confidence = "✅ High" if score >= FUZZY_THRESHOLD else "⚠️ Low"
        
#         rows.append({
#             "image_filename": filename,
#             "gt_name": gt_name,
#             "pred_name": pred_item["name"],
#             "similarity_score": round(score, 1),
#             "gt_qty": gt_qty,
#             "pred_qty": pred_qty,
#             "diff": round(diff, 2),
#             "abs_diff": round(abs_diff, 2),
#             "pct_diff": round(pct_diff, 1) if pct_diff is not None else None,
#             "confidence": confidence,
#             "source": pred_item.get('source', 'unknown')
#         })
    
#     # Process unmatched ground truth ingredients
#     for gt_idx, gt_item in enumerate(gt_ings):
#         if gt_idx not in matched_gt_indices:
#             gt_name = gt_item["name"]
#             gt_qty = float(gt_item.get("quantity", 0))
            
#             rows.append({
#                 "image_filename": filename,
#                 "gt_name": gt_name,
#                 "pred_name": None,
#                 "similarity_score": 0,
#                 "gt_qty": gt_qty,
#                 "pred_qty": None,
#                 "diff": None,
#                 "abs_diff": None,
#                 "pct_diff": None,
#                 "confidence": "❌ No Match",
#                 "source": None
#             })
    
#     # Process unmatched predicted ingredients
#     for pred_idx, pred_item in enumerate(all_pred):
#         if pred_idx not in matched_pred_indices:
#             rows.append({
#                 "image_filename": filename,
#                 "gt_name": None,
#                 "pred_name": pred_item["name"],
#                 "similarity_score": 0,
#                 "gt_qty": None,
#                 "pred_qty": float(pred_item.get("quantity", 0)),
#                 "diff": None,
#                 "abs_diff": None,
#                 "pct_diff": None,
#                 "confidence": "🔍 Extra Prediction",
#                 "source": pred_item.get('source', 'unknown')
#             })

# # ====== Export to Excel ======
# df = pd.DataFrame(rows)
# df.to_excel(OUTPUT_XLSX, index=False)

# # ====== Open Excel and Merge Cells with Same Image Name ======
# wb = load_workbook(OUTPUT_XLSX)
# ws = wb.active

# # Auto-adjust column widths
# for col in ws.columns:
#     max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
#     ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# # Merge cells with same filename
# col_idx = 1  # First column
# start_row = 2
# for i in range(2, ws.max_row + 2):
#     if i == ws.max_row + 1 or ws.cell(i, col_idx).value != ws.cell(start_row, col_idx).value:
#         if i - start_row > 1:
#             ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=i-1, end_column=col_idx)
#             ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
#         start_row = i

# wb.save(OUTPUT_XLSX)
# wb.close()

# print(f"✅ Completed! Generated {len(df)} rows, results saved to: {OUTPUT_XLSX}")
