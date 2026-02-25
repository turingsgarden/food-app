import openpyxl
from openpyxl import load_workbook, Workbook
from statistics import mean

# === Input and Output Paths ===
input_path = "Evaluation/Nutrition5k_100-300/ingredient_comparison_optimaized_100-300.xlsx" or "Evaluation/Nutrition5k_700-800/ingredient_comparison_optimaized_700-800.xlsx"
output_path = "Evaluation/Nutrition5k_100-300/ingredient_comparison_filtered_merged.xlsx" or "Evaluation/Nutrition5k_700-800/ingredient_comparison_filtered_merged_700-800.xlsx"




# === Read Original File ===
wb = load_workbook(input_path)
ws = wb.active

# === Create Output File ===
new_wb = Workbook()
new_ws = new_wb.active
headers = [
    "image_filename", "dish difference", "avg_diff", "avg_pct_diff",
    "gt_name", "pred_name", "similarity_score", "gt_qty", "pred_qty",
    "diff", "abs_diff", "pct_diff", "confidence"
]
new_ws.append(headers)

# === Get All Merged Cell Information ===
merged_ranges = ws.merged_cells.ranges

def find_merge_range(cell):
    """Determine if the cell is in a merged range, return the merged range object"""
    for merged in merged_ranges:
        if cell.coordinate in merged:
            return merged
    return None

row = 2
max_row = ws.max_row

# === Collect All Blocks to Write ===
blocks = []

print("=== Starting Data Processing ===")

while row <= max_row:
    img_cell = ws.cell(row, 1)
    diff_cell = ws.cell(row, 2)
    img_merge = find_merge_range(img_cell)

    if img_merge:
        start_row, end_row = img_merge.min_row, img_merge.max_row
        dish_diff_value = ws.cell(start_row, 2).value
        img_value = ws.cell(start_row, 1).value
    else:
        start_row = end_row = row
        dish_diff_value = diff_cell.value
        img_value = img_cell.value

    print(f"\nProcessing image: {img_value}, row range: {start_row}-{end_row}, dish_difference: {dish_diff_value}")

    # Skip blocks with "No"
    if str(dish_diff_value).strip().lower() == "no":
        print(f"Skipping No block: {img_value}")
        row = end_row + 1
        continue

    # Calculate averages - corrected column numbers!
    diffs, pct_diffs = [], []
    print(f"Calculating averages - scanning rows {start_row} to {end_row}:")
    
    for r in range(start_row, end_row + 1):
        diff = ws.cell(r, 9).value        # diff in column 9 (correct)
        pct_diff = ws.cell(r, 10).value   # pct_diff in column 10 (corrected here!)
        
        # Print each cell value
        print(f"  Row {r}: diff={diff} (col 9), pct_diff={pct_diff} (col 10)")
        
        try:
            if diff is not None:
                diffs.append(float(diff))
                print(f"    -> Added diff: {float(diff)}")
            if pct_diff is not None:
                pct_diffs.append(float(pct_diff))
                print(f"    -> Added pct_diff: {float(pct_diff)}")
        except Exception as e:
            print(f"    -> Conversion error: {e}")
            continue

    avg_diff = round(mean(diffs), 3) if diffs else ""
    avg_pct_diff = round(mean(pct_diffs), 3) if pct_diffs else ""
    
    print(f"Calculation results: diffs list={diffs}, pct_diffs list={pct_diffs}")
    print(f"Averages: avg_diff={avg_diff}, avg_pct_diff={avg_pct_diff}")

    block_rows = []
    for r in range(start_row, end_row + 1):
        row_data = [
            img_value,
            dish_diff_value,
            avg_diff,
            avg_pct_diff,
            ws.cell(r, 3).value,   # gt_name (column 3)
            ws.cell(r, 4).value,   # pred_name (column 4)
            ws.cell(r, 5).value,   # similarity_score (column 5)
            ws.cell(r, 6).value,   # gt_qty (column 6)
            ws.cell(r, 7).value,   # pred_qty (column 7)
            ws.cell(r, 8).value,   # diff (column 8)
            ws.cell(r, 9).value,   # abs_diff (column 9)
            ws.cell(r, 10).value,  # pct_diff (column 10)
            ws.cell(r, 11).value,  # confidence (column 11)
        ]
        block_rows.append(row_data)

    blocks.append((block_rows, len(block_rows)))
    row = end_row + 1

# === Write to New Sheet + Merge First Four Columns ===
current_row = 2
print(f"\n=== Starting Data Writing, Total {len(blocks)} blocks ===")

for i, (block_rows, num_rows) in enumerate(blocks):
    print(f"Writing block {i+1}: {num_rows} rows")
    for j, row_data in enumerate(block_rows):
        new_ws.append(row_data)
        print(f"  Writing row {current_row + j}: avg_diff={row_data[2]}, avg_pct_diff={row_data[3]}")

    if num_rows > 1:
        for col in range(1, 5):  # Merge first four columns
            new_ws.merge_cells(
                start_row=current_row,
                start_column=col,
                end_row=current_row + num_rows - 1,
                end_column=col
            )
        print(f"  Merged first 4 columns: rows {current_row} to {current_row + num_rows - 1}")
    
    current_row += num_rows

# === Save ===
new_wb.save(output_path)
print(f"\n✅ Processing completed! File saved to: {output_path}")
print(f"Total processed {len(blocks)} image blocks")