# import os, shutil, random

# SRC = "/home/sheru/.cache/kagglehub/datasets/rkuo2000/uecfood256/versions/1"
# DST = os.path.expanduser("~/datasets/uecfood256-small-random")
# os.makedirs(DST, exist_ok=True)

# # Collect all image paths
# exts = (".jpg", ".jpeg", ".png")
# all_images = []
# for dirpath, _, files in os.walk(SRC):
#     for fn in files:
#         if fn.lower().endswith(exts):
#             all_images.append(os.path.join(dirpath, fn))

# random.shuffle(all_images)

# total_size = 0
# limit = 500 * 1024 * 1024

# for img_path in all_images:
#     rel = os.path.relpath(img_path, SRC)
#     dst_path = os.path.join(DST, rel)
#     os.makedirs(os.path.dirname(dst_path), exist_ok=True)
#     shutil.copy2(img_path, dst_path)
#     total_size += os.path.getsize(dst_path)
#     if total_size >= limit:
#         print(f"✅ Random subset ready ({total_size/1e6:.1f} MB)")
#         break

#above script was to take a random selection of the large ~3gb, below is to add images in csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os

CSV_PATH = "analysis_results.csv"
XLSX_PATH = "analysis_results_with_images.xlsx"

# Step 1: Load CSV
# df = pd.read_csv(CSV_PATH, sep=",|\t", engine="python")  # supports tab or comma
df = pd.read_csv(CSV_PATH)
print(f"Loaded {len(df)} rows")

# Step 2: Write to Excel first
df.to_excel(XLSX_PATH, index=False)
wb = load_workbook(XLSX_PATH)
ws = wb.active

# Find column indexes
headers = [cell.value for cell in ws[1]]
img_col_idx = headers.index("image_path") + 1  # openpyxl is 1-based
insert_col_idx = img_col_idx + 1  # add image right next to path

# Shift other columns to the right of inserted column
ws.insert_cols(insert_col_idx)
ws.cell(row=1, column=insert_col_idx, value="Thumbnail")

# Step 3: Embed image thumbnails
for row in range(2, ws.max_row + 1):
    img_path = ws.cell(row=row, column=img_col_idx).value
    if not os.path.exists(img_path):
        continue
    try:
        img = XLImage(img_path)
        img.width, img.height = 80, 80  # small thumbnails
        cell_ref = f"{get_column_letter(insert_col_idx)}{row}"
        ws.add_image(img, cell_ref)
        ws.row_dimensions[row].height = 65  # adjust row height
    except Exception as e:
        print(f"Skipping {img_path}: {e}")

# Step 4: Adjust column widths
for col in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col)].width = 30

wb.save(XLSX_PATH)
print(f"✅ Saved Excel file with thumbnails → {XLSX_PATH}")