
"""
Strict mass evaluation script
Match predicted food mass to ground truth mass using improved fuzzy matching,
use both visible and hidden ingredients for matching,
excludes certain ground truth entries based on keywords to avoid ambiguity
"deprecated", "Mixed green", "plate only", "salad"
and generate a detailed Excel report comparing mass quantities
"""
import json
import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re

# ====== File Paths ======
PRED_PATH = "model_output_analysis/mass_prediction_data_and_result/mass_prediction.json"
GROUND_TRUTH_PATH = "model-answer-display/Nutrition5k/metadata/dish_metadata_cafe1.json"
OUTPUT_XLSX = "model_output_analysis/mass_prediction_data_and_result/mass_comparison.xlsx"
# ====== Fuzzy Matching Threshold ======
FUZZY_THRESHOLD = 80

# ====== Read JSON Files ======
with open(GROUND_TRUTH_PATH, "r", encoding="utf-8") as f:
    gt_data = json.load(f)
with open(PRED_PATH, "r", encoding="utf-8") as f:
    pred_data = json.load(f)

# ====== Build Prediction Lookup Table ======
pred_dict = {p["dish_id"]: p for p in pred_data}

# ====== Extract dish_id from filename ======
def extract_dish_id_from_filename(filename):

    if not filename:
        return None

    match = re.search(r'dish_(\d+)_', filename)
    return match.group(1) if match else None

# ====== Filter Function ======
def should_skip_ground_truth(gt_entry) -> bool:
    """Check if this ground truth entry should be skipped"""
    skip_keywords = ["deprecated", "Mixed green", "plate only", "salad"]
    
    # Check if ingredients contain any skip keywords
    ingredients = gt_entry.get("ingredients", [])
    for ingredient in ingredients:
        ingredient_name = ingredient.get("name", "").lower()
        for keyword in skip_keywords:
            if keyword.lower() in ingredient_name:
                return True
    
    return False

# ====== Enhanced Normalization Function ======
def normalize_name(name: str) -> str:
    """Enhanced text normalization"""
    if not name:
        return ""
    name = name.lower().strip()
    name = re.sub(r"[^a-z\s]", "", name)  # Remove symbols
    
    # Handle common plurals and synonyms
    plural_map = {
        "berries": "berry", "grapes": "grape", "olives": "olive",
        "tomatoes": "tomato", "potatoes": "potato",
        "oranges": "orange", "tangerines": "tangerine",
        "mandarins": "mandarin"
    }
    for plural, singular in plural_map.items():
        if plural in name:
            name = name.replace(plural, singular)
    
    # Handle specific synonyms
    synonym_map = {
        "mandarin orange": "tangerine",
        "mandarins": "tangerine",
        "mandarin": "tangerine",
        "tangerines": "tangerine",
        "tangreines": "tangerine",  # Handle typo
        "tangenine": "tangerine",   # Handle typo
        "tagenrine": "tangerine"    # Handle typo
    }
    
    for synonym, standard in synonym_map.items():
        if synonym in name:
            name = name.replace(synonym, standard)
    
    name = " ".join(name.split())  # Normalize whitespace
    return name

# ====== Check for Mandarin-Tangerine Relationship ======
def is_mandarin_tangerine_match(gt_name: str, pred_name: str) -> bool:
    """Check if names refer to the same fruit (mandarin oranges and tangerines)"""
    if not gt_name or not pred_name:
        return False
    
    gt_lower = gt_name.lower()
    pred_lower = pred_name.lower()
    
    # Define all possible variations
    mandarin_variations = ["mandarin orange", "mandarins", "mandarin", "mandarin oranges"]
    tangerine_variations = ["tangerine", "tangerines", "tagenrine", "tangreines", "tangenine"]
    
    # Check if one is mandarin and the other is tangerine
    is_gt_mandarin = any(variation in gt_lower for variation in mandarin_variations)
    is_pred_mandarin = any(variation in pred_lower for variation in mandarin_variations)
    
    is_gt_tangerine = any(variation in gt_lower for variation in tangerine_variations)
    is_pred_tangerine = any(variation in pred_lower for variation in tangerine_variations)
    
    # They match if (gt is mandarin and pred is tangerine) OR (gt is tangerine and pred is mandarin)
    return (is_gt_mandarin and is_pred_tangerine) or (is_gt_tangerine and is_pred_mandarin)

# ====== Fuzzy Matching Function ======
def fuzzy_match(gt_name, pred_names):
    """Return the most similar predicted food name and score"""
    if not pred_names:
        return None, 0
    match, score, _ = process.extractOne(gt_name, pred_names, scorer=fuzz.token_sort_ratio)
    return match, score

# ====== Find Best Matching Predictions ======
def find_best_matching_predictions(gt_name, pred_food_items):
    """Find best matching predicted food items based on name similarity, even if low score"""
    normalized_gt = normalize_name(gt_name)
    
    if not pred_food_items:
        return [], 0, None, "❌ No Match"
    
    pred_names = [normalize_name(p["name"]) for p in pred_food_items]
    
    # 0️⃣ First check for mandarin-tangerine special case
    for pred_item in pred_food_items:
        pred_original_name = pred_item["name"]
        if is_mandarin_tangerine_match(gt_name, pred_original_name):
            # Special case: mandarin oranges and tangerines are considered the same
            matching_items = [pred_item]
            best_original_name = pred_original_name
            # Give it a high score for this special case
            score = 95  # High score for this special match
            return matching_items, score, best_original_name, "✅ High"
    
    # 1️⃣ First try fuzzy matching to find the best match
    match_name, score = fuzzy_match(normalized_gt, pred_names)
    
    if not match_name:
        return [], 0, None, "❌ No Match"
    
    # Find all items with the same normalized name as the best match
    matching_items = [p for p in pred_food_items if normalize_name(p["name"]) == match_name]
    
    # Get original names for display
    if matching_items:
        best_original_names = [p["name"] for p in matching_items]
        best_original_name = " + ".join(best_original_names)
    else:
        # If no exact match, try to find the item with the best matching name
        best_original_name = None
        for pred_item in pred_food_items:
            if normalize_name(pred_item["name"]) == match_name:
                best_original_name = pred_item["name"]
                matching_items = [pred_item]
                break
    
    # Determine confidence level
    if score >= FUZZY_THRESHOLD:
        confidence = "✅ High"
    elif score > 0:
        # Check containment relationship
        for pred_item in pred_food_items:
            pred_norm = normalize_name(pred_item["name"])
            if (normalized_gt in pred_norm or pred_norm in normalized_gt):
                confidence = "🟡 Contained"
                break
        else:
            confidence = f"❌ Low Score ({score:.1f})"
    else:
        confidence = "❌ No Match"
    
    return matching_items, score, best_original_name, confidence

# ====== Determine Prediction Correctness ======
def determine_prediction_correctness(dish_rows):
    """Determine if the entire dish prediction is correct based on confidence values"""
    for row in dish_rows:
        if row.get("confidence") and row["confidence"].startswith("❌"):
            return "Incorrect"
    return "Correct"

# ====== Main Loop ======
rows = []
skipped_count = 0
processed_count = 0
not_found_count = 0
mandarin_tangerine_matches = 0  # Counter for special matches

# First, collect all dish data to determine correctness
dish_data = {}

for dish_id, pred_entry in pred_dict.items():
    
    gt_entry = None
    for g in gt_data:
        gt_dish_id = extract_dish_id_from_filename(g.get("image_filename"))
        if gt_dish_id == dish_id:
            gt_entry = g
            break
    
    if not gt_entry:
        print(f"⚠️ No ground truth found for dish_id: {dish_id}")
        not_found_count += 1
        continue
    
    # Check if ground truth contains skip keywords
    if should_skip_ground_truth(gt_entry):
        skipped_count += 1
        continue
    
    processed_count += 1
    
    # Get predicted food items and total mass
    pred_food_items = pred_entry.get("mass_estimation", {}).get("food_items", [])
    pred_total_mass = pred_entry.get("mass_estimation", {}).get("total_mass_g", 0)
    
    # Get ground truth ingredients
    gt_ings = gt_entry.get("ingredients", [])
    
    # Calculate total ground truth mass
    gt_total_mass = sum(float(ing.get("quantity", 0)) for ing in gt_ings)
    
    # Calculate total differences
    total_diff = pred_total_mass - gt_total_mass
    total_abs_diff = abs(total_diff)
    total_pct_diff = (total_diff / gt_total_mass * 100) if gt_total_mass != 0 else None
    
    # Process individual food items
    ingredient_data = []
    for gt_item in gt_ings:
        gt_name = gt_item["name"]
        gt_qty = float(gt_item.get("quantity", 0))

        # Use matching logic to find best match, even if low score
        matching_items, score, best_match_name, confidence = find_best_matching_predictions(gt_name, pred_food_items)

        # Count mandarin-tangerine special matches
        if confidence == "✅ High" and is_mandarin_tangerine_match(gt_name, best_match_name if best_match_name else ""):
            mandarin_tangerine_matches += 1

        # ALWAYS calculate prediction quantity if we have matching items
        total_pred_qty = None
        if matching_items:
            total_pred_qty = sum(float(p.get("predicted_mass_g", 0)) for p in matching_items)
        
        # ALWAYS calculate differences, even for mismatches
        ingredient_diff = None
        ingredient_abs_diff = None
        ingredient_pct_diff = None
        
        if total_pred_qty is not None and gt_qty > 0:
            ingredient_diff = total_pred_qty - gt_qty
            ingredient_abs_diff = abs(ingredient_diff)
            ingredient_pct_diff = (ingredient_diff / gt_qty * 100)
        
        ingredient_data.append({
            "ingredient_gt_name": gt_name,
            "ingredient_pred_name": best_match_name,  # Always show best match
            "similarity_score": round(score, 1) if score is not None else None,
            "ingredient_gt_qty": gt_qty,
            "ingredient_pred_qty": round(total_pred_qty, 2) if total_pred_qty is not None else None,
            "ingredient_diff": round(ingredient_diff, 2) if ingredient_diff is not None else None,
            "ingredient_abs_diff": round(ingredient_abs_diff, 2) if ingredient_abs_diff is not None else None,
            "ingredient_pct_diff": round(ingredient_pct_diff, 1) if ingredient_pct_diff is not None else None,
            "confidence": confidence
        })
    
    # Calculate average ingredient differences (for all ingredients with predictions)
    ingredients_with_predictions = [row for row in ingredient_data if row["ingredient_pred_qty"] is not None]
    if ingredients_with_predictions:
        total_avg_diff = sum(row["ingredient_abs_diff"] for row in ingredients_with_predictions 
                            if row["ingredient_abs_diff"] is not None) / len(ingredients_with_predictions)
    else:
        total_avg_diff = None
    
    # Store dish data for correctness determination
    dish_data[dish_id] = {
        "base_row": {
            "dish_id": dish_id,
            "gt_total_mass": round(gt_total_mass, 2),
            "pred_total_mass": round(pred_total_mass, 2),
            "total_diff": round(total_diff, 2),
            "total_abs_diff": round(total_abs_diff, 2),
            "total_pct_diff": round(total_pct_diff, 1) if total_pct_diff is not None else None,
            "total_avg_diff": round(total_avg_diff, 2) if total_avg_diff is not None else None,
            "scale_factor_used": pred_entry.get("mass_estimation", {}).get("scale_factor_used", 1.0)
        },
        "ingredient_data": ingredient_data
    }

# Now build the final rows with prediction correctness
for dish_id, data in dish_data.items():
    base_row = data["base_row"]
    ingredient_data = data["ingredient_data"]
    
    # Determine prediction correctness for this dish
    prediction_correctness = determine_prediction_correctness(ingredient_data)
    
    # Add prediction correctness to base row
    base_row["prediction_correctness"] = prediction_correctness
    
    # Create separate rows for each ingredient
    if ingredient_data:
        # First ingredient row includes dish summary
        first_ingredient = ingredient_data[0]
        row = base_row.copy()
        row.update({
            "ingredient_gt_name": first_ingredient["ingredient_gt_name"],
            "ingredient_pred_name": first_ingredient["ingredient_pred_name"],
            "similarity_score": first_ingredient["similarity_score"],
            "ingredient_gt_qty": first_ingredient["ingredient_gt_qty"],
            "ingredient_pred_qty": first_ingredient["ingredient_pred_qty"],
            "ingredient_diff": first_ingredient["ingredient_diff"],
            "ingredient_abs_diff": first_ingredient["ingredient_abs_diff"],
            "ingredient_pct_diff": first_ingredient["ingredient_pct_diff"],
            "confidence": first_ingredient["confidence"]
        })
        rows.append(row)
        
        # Remaining ingredient rows (dish summary columns will be empty)
        for ingredient in ingredient_data[1:]:
            row = {
                "dish_id": dish_id,
                "gt_total_mass": None,
                "pred_total_mass": None,
                "total_diff": None,
                "total_abs_diff": None,
                "total_pct_diff": None,
                "total_avg_diff": None,
                "scale_factor_used": None,
                "prediction_correctness": None,  # Will be merged in Excel
                "ingredient_gt_name": ingredient["ingredient_gt_name"],
                "ingredient_pred_name": ingredient["ingredient_pred_name"],
                "similarity_score": ingredient["similarity_score"],
                "ingredient_gt_qty": ingredient["ingredient_gt_qty"],
                "ingredient_pred_qty": ingredient["ingredient_pred_qty"],
                "ingredient_diff": ingredient["ingredient_diff"],
                "ingredient_abs_diff": ingredient["ingredient_abs_diff"],
                "ingredient_pct_diff": ingredient["ingredient_pct_diff"],
                "confidence": ingredient["confidence"]
            }
            rows.append(row)
    else:
        # No ingredients case
        row = base_row.copy()
        row.update({
            "ingredient_gt_name": None,
            "ingredient_pred_name": None,
            "similarity_score": None,
            "ingredient_gt_qty": None,
            "ingredient_pred_qty": None,
            "ingredient_diff": None,
            "ingredient_abs_diff": None,
            "ingredient_pct_diff": None,
            "confidence": None
        })
        rows.append(row)

# ====== Export to Excel ======
if rows:
    df = pd.DataFrame(rows)
    
    # Reorder columns to match desired structure with prediction_correctness added
    column_order = [
        'dish_id', 'gt_total_mass', 'pred_total_mass', 'total_diff', 
        'total_abs_diff', 'total_pct_diff', 'total_avg_diff', 'scale_factor_used',
        'prediction_correctness',  # New column
        'ingredient_gt_name', 'ingredient_pred_name', 'similarity_score',
        'ingredient_gt_qty', 'ingredient_pred_qty', 'ingredient_diff', 
        'ingredient_abs_diff', 'ingredient_pct_diff', 'confidence'
    ]
    df = df[column_order]
    
    df.to_excel(OUTPUT_XLSX, index=False)

    # ====== Excel Formatting ======
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Merge cells for dish summary information including prediction_correctness
    current_dish = None
    start_row = 2  # Start from row 2 (after header)
    end_row = 2
    
    for row_idx in range(2, ws.max_row + 2):  # +2 to include the last row
        dish_id_cell = ws.cell(row=row_idx, column=1)
        
        if row_idx == ws.max_row + 1 or dish_id_cell.value != current_dish:
            if current_dish is not None and end_row - start_row > 0:
                # Merge dish summary columns for this dish (columns A-I: dish_id to prediction_correctness)
                for col_idx in range(1, 10):  # Columns A-I (dish_id to prediction_correctness)
                    if start_row < end_row:  # Only merge if there are multiple rows
                        ws.merge_cells(
                            start_row=start_row, start_column=col_idx,
                            end_row=end_row, end_column=col_idx
                        )
                        # Center align the merged cell
                        ws.cell(start_row, col_idx).alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
            
            if row_idx <= ws.max_row:
                current_dish = dish_id_cell.value
                start_row = row_idx
                end_row = row_idx
        else:
            end_row = row_idx

    # Add styling for better readability
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
        for cell in row:
            if cell.row == 1:  # Header row
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif cell.column in [2, 3, 4, 5, 6, 7, 8, 12, 13, 14, 15, 16]:  # Numeric columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:  # Text columns
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Add conditional formatting for confidence column
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    
    # Define fills for different confidence levels
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    
    # Get the column index for confidence (last column)
    confidence_col = len(df.columns)
    
    # Apply conditional formatting
    for row in range(2, ws.max_row + 1):
        confidence_value = ws.cell(row=row, column=confidence_col).value
        if confidence_value:
            if "✅ High" in confidence_value:
                ws.cell(row=row, column=confidence_col).fill = high_fill
            elif "🟡 Contained" in confidence_value:
                ws.cell(row=row, column=confidence_col).fill = medium_fill
            elif "❌" in confidence_value:
                ws.cell(row=row, column=confidence_col).fill = low_fill

    wb.save(OUTPUT_XLSX)
    wb.close()

    print(f"✅ Completed! Processed {processed_count} dishes, skipped {skipped_count} dishes containing filter keywords, {not_found_count} dishes not found in ground truth")
    print(f"📊 Generated {len(df)} rows, saved to: {OUTPUT_XLSX}")

    # ====== Generate Summary Statistics ======
    print(f"\n📈 Summary Statistics:")
    
    # Total mass statistics (only use first row of each dish to avoid duplicates)
    dish_summary_rows = [row for row in rows if row['gt_total_mass'] is not None]
    if dish_summary_rows:
        total_mae = sum(row['total_abs_diff'] for row in dish_summary_rows if row['total_abs_diff'] is not None) / len(dish_summary_rows)
        total_pct_diffs = [row['total_pct_diff'] for row in dish_summary_rows if row['total_pct_diff'] is not None]
        total_avg_pct_diff = sum(total_pct_diffs) / len(total_pct_diffs) if total_pct_diffs else None
        
        print(f"   Total Mass MAE: {total_mae:.2f}g")
        if total_avg_pct_diff is not None:
            print(f"   Average Total Mass Percentage Difference: {total_avg_pct_diff:.1f}%")
    
    # Ingredient matching statistics
    total_ingredients = len([row for row in rows if row['ingredient_gt_name'] is not None])
    high_matches = len([row for row in rows if row['confidence'] and '✅ High' in row['confidence']])
    medium_matches = len([row for row in rows if row['confidence'] and '🟡 Contained' in row['confidence']])
    low_matches = len([row for row in rows if row['confidence'] and '❌ Low Score' in row['confidence']])
    no_matches = len([row for row in rows if row['confidence'] and '❌ No Match' in row['confidence']])
    
    if total_ingredients > 0:
        print(f"   Ingredient matching breakdown:")
        print(f"     ✅ High matches: {high_matches} ({high_matches/total_ingredients*100:.1f}%)")
        print(f"     🟡 Contained matches: {medium_matches} ({medium_matches/total_ingredients*100:.1f}%)")
        print(f"     ❌ Low Score matches: {low_matches} ({low_matches/total_ingredients*100:.1f}%)")
        print(f"     ❌ No Match: {no_matches} ({no_matches/total_ingredients*100:.1f}%)")
    
    # Prediction correctness statistics
    correct_dishes = len([row for row in dish_summary_rows if row.get('prediction_correctness') == 'Correct'])
    total_dishes = len(dish_summary_rows)
    if total_dishes > 0:
        correctness_rate = (correct_dishes / total_dishes) * 100
        print(f"   Dish prediction correctness rate: {correctness_rate:.1f}% ({correct_dishes}/{total_dishes})")
    
    # Scale factor statistics
    unique_scale_factors = set(row['scale_factor_used'] for row in dish_summary_rows if row['scale_factor_used'] is not None)
    print(f"   Scale factors used: {', '.join(map(str, unique_scale_factors))}")
    
    # Mandarin-tangerine special matches
    if mandarin_tangerine_matches > 0:
        print(f"   Mandarin-Tangerine special matches: {mandarin_tangerine_matches}")
    
    # Ingredients with predictions (even if wrong)
    ingredients_with_pred = len([row for row in rows if row['ingredient_pred_qty'] is not None])
    print(f"   Ingredients with prediction data: {ingredients_with_pred}/{total_ingredients} ({ingredients_with_pred/total_ingredients*100:.1f}%)")
    
else:

    print("❌ No data to export! Check if dish_id matching is working correctly.")
