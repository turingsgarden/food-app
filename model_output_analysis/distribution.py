import openpyxl
import matplotlib.pyplot as plt
from statistics import mean
import numpy as np


input_path = "model_output_analysis/mass_prediction_data_and_result/mass_comparison.xlsx"


COLUMN_DISH_ID = 1                    # colA: dish_id
COLUMN_TOTAL_PCT_DIFF = 6             # colF: total_pct_diff
COLUMN_PREDICTION_CORRECTNESS = 9     # colI: prediction_correctness
COLUMN_INGREDIENT_PCT_DIFF = 17       # colQ: ingredient_pct_diff
COLUMN_CONFIDENCE = 18                # colR: confidence

# === Loading file ===
wb = openpyxl.load_workbook(input_path, data_only=True)
ws = wb.active

# === Collecting data ===
dish_pct_diffs = []      
ingredient_pct_diffs = [] 
correct_dish_ids = set() 

# collecting correct dish total_pct_diff
last_dish_id = None
for row in range(2, ws.max_row + 1):
    dish_id = ws.cell(row, COLUMN_DISH_ID).value
    correctness = ws.cell(row, COLUMN_PREDICTION_CORRECTNESS).value
    total_pct_diff = ws.cell(row, COLUMN_TOTAL_PCT_DIFF).value
    
   
    if correctness == "Correct" and dish_id is not None and dish_id != last_dish_id:
        try:
            if total_pct_diff is not None and total_pct_diff != "":
                dish_pct_diffs.append({
                    'value': float(total_pct_diff),
                    'dish_id': dish_id,
                    'row': row
                })
                correct_dish_ids.add(dish_id)
        except ValueError:
            pass
        last_dish_id = dish_id

#collecting all correct dish的ingredient_pct_diff
current_dish_id = None
for row in range(2, ws.max_row + 1):
    dish_id = ws.cell(row, COLUMN_DISH_ID).value
    if dish_id is not None:
        current_dish_id = dish_id
    
    ingredient_pct_diff = ws.cell(row, COLUMN_INGREDIENT_PCT_DIFF).value
    confidence = ws.cell(row, COLUMN_CONFIDENCE).value
    
   
    if (confidence in ["✅ High", "🟡 Contained"] and 
        ingredient_pct_diff is not None and 
        ingredient_pct_diff != ""):
        try:
            ingredient_pct_diffs.append({
                'value': float(ingredient_pct_diff),
                'dish_id': current_dish_id,
                'row': row,
                'confidence': confidence
            })
        except ValueError:
            continue

print("=== Data Summary ===")
print(f"Correct dishes found: {len(correct_dish_ids)}")
print(f"Dish total_pct_diff samples: {len(dish_pct_diffs)}")
print(f"Ingredient pct_diff samples (name prediction correct): {len(ingredient_pct_diffs)}")


if ingredient_pct_diffs:
    high_conf = sum(1 for item in ingredient_pct_diffs if item.get('confidence') == "✅ High")
    contained_conf = sum(1 for item in ingredient_pct_diffs if item.get('confidence') == "🟡 Contained")
    print(f"  - ✅ High confidence: {high_conf} ingredients")
    print(f"  - 🟡 Contained confidence: {contained_conf} ingredients")
    
    
    unique_dishes = set(item['dish_id'] for item in ingredient_pct_diffs)
    print(f"  - Unique dishes in ingredient analysis: {len(unique_dishes)}")

# === Helper function: Find min and max values with dish_id ===
def find_min_max_with_ids(data_list):
    """Find the minimum and maximum values and their corresponding dish_id from the data list"""
    if not data_list:
        return None, None, None, None
    
    min_val = min(item['value'] for item in data_list)
    max_val = max(item['value'] for item in data_list)
    
 
    min_dish_ids = [item['dish_id'] for item in data_list if abs(item['value'] - min_val) < 1e-6]
    max_dish_ids = [item['dish_id'] for item in data_list if abs(item['value'] - max_val) < 1e-6]
    
    min_dish_id_str = min_dish_ids[0] if len(min_dish_ids) == 1 else f"{min_dish_ids[0]}等{len(min_dish_ids)}个"
    max_dish_id_str = max_dish_ids[0] if len(max_dish_ids) == 1 else f"{max_dish_ids[0]}等{len(max_dish_ids)}个"
    
    return min_val, max_val, min_dish_id_str, max_dish_id_str

# === Helper function: Create uniform width bins ===
def create_uniform_bins(data_values):
    
    if len(data_values) == 0:
        return [], [], []
    
    # Extract values
    values = np.array(data_values)
    min_val = min(values)
    max_val = max(values)
    
    # Collect all bin boundaries and labels
    bin_ranges = []  # Actual range of each bin (start, end)
    bin_labels = []   # Display label for each bin
    x_positions = []  # Position of each bin on the x-axis (starting from 0)
    
    # 1. Handle extreme negative values (< -1000%)
    if min_val < -1000:
        
        extreme_neg_start = np.floor(min_val / 100) * 100  # Round down to nearest 100%
        extreme_neg_end = -1000
        bin_ranges.append((extreme_neg_start, extreme_neg_end))
        bin_labels.append(f"< -1000%\n[{int(extreme_neg_start)}%, -1000%]")
        x_positions.append(len(bin_ranges) - 1)
    
    # 2. Handle negative outer region (-1000% to -200%)
    if min_val < -200:
        neg_outer_start = max(-1000, np.floor(min_val / 100) * 100)
        neg_outer_end = -200
        
        # Create bins of 100% each
        current = neg_outer_start
        while current < neg_outer_end:
            bin_end = min(current + 100, neg_outer_end)
            bin_ranges.append((current, bin_end))
            bin_labels.append(f"[{int(current)}%, {int(bin_end)}%]")
            x_positions.append(len(bin_ranges) - 1)
            current = bin_end
    
    # 3. Center region (-200% to 200%)
    center_start = -200
    center_end = 200
    center_bin_width = 10
    
    # Create bins of 10% each
    current = center_start
    while current < center_end:
        bin_end = min(current + center_bin_width, center_end)
        bin_ranges.append((current, bin_end))
        
        # Simplify labels: show detailed labels every 50%, otherwise only show the start value
        if current % 50 == 0 or current == -200 or current == 200 - center_bin_width:
            bin_labels.append(f"[{int(current)}%, {int(bin_end)}%]")
        else:
            bin_labels.append(f"{int(current)}%")
        
        x_positions.append(len(bin_ranges) - 1)
        current = bin_end
    
    # 4. Handle positive outer region (200% to 1000%)
    if max_val > 200:
        pos_outer_start = 200
        pos_outer_end = min(1000, np.ceil(max_val / 100) * 100)
        
        # Create bins of 100% each
        current = pos_outer_start
        while current < pos_outer_end:
            bin_end = min(current + 100, pos_outer_end)
            bin_ranges.append((current, bin_end))
            bin_labels.append(f"[{int(current)}%, {int(bin_end)}%]")
            x_positions.append(len(bin_ranges) - 1)
            current = bin_end
    
    # 5. Handle extreme positive values (> 1000%)
    if max_val > 1000:
        # Merge all >1000% into one bin
        extreme_pos_start = 1000
        extreme_pos_end = np.ceil(max_val / 100) * 100  # Round up to nearest 100%
        bin_ranges.append((extreme_pos_start, extreme_pos_end))
        bin_labels.append(f"> 1000%\n[1000%, {int(extreme_pos_end)}%]")
        x_positions.append(len(bin_ranges) - 1)
    
    return x_positions, bin_ranges, bin_labels

# === Helper function: Assign data to bins by category ===
def assign_data_to_bins_by_category(values, bin_ranges):
    """
    Assign data to bins by category, returning counts of three categories for each bin
    Categories: Underestimate (<0), Accurate (0), Overestimate (>0)
    """
    bin_counts_under = [0] * len(bin_ranges)  
    bin_counts_accurate = [0] * len(bin_ranges)  
    bin_counts_over = [0] * len(bin_ranges)  
    
    for val in values:
        # Determine category
        if abs(val) < 0.1:  # Accurate prediction (within ±0.1%)
            category = 'accurate'
        elif val < 0:       # Underestimate
            category = 'under'
        else:               # Overestimate
            category = 'over'
        
        # Assign to corresponding bin
        for i, (bin_start, bin_end) in enumerate(bin_ranges):
            if bin_start <= val < bin_end or (i == len(bin_ranges) - 1 and val == bin_end):
                if category == 'under':
                    bin_counts_under[i] += 1
                elif category == 'accurate':
                    bin_counts_accurate[i] += 1
                else:  # 'over'
                    bin_counts_over[i] += 1
                break
    
    return bin_counts_under, bin_counts_accurate, bin_counts_over

# === Helper function: Plot stacked histogram (green in the middle) ===
def plot_stacked_histogram_green_in_middle(ax, data_values, title, xlabel, data_list=None, is_ingredient=False):
    """
    Plot stacked histogram with green in the middle between blue and red
    Stacking order: Blue (underestimate) at bottom, Green (accurate) in middle, Red (overestimate) at top
    """
    if len(data_values) == 0:
        return None, None, None, None
    
    values = np.array(data_values)
    

    min_val = min(values)
    max_val = max(values)
    mean_val = round(mean(values), 3)
    count = len(values)
    
    if is_ingredient:
        print(f"\n=== Ingredient-Level Analysis (Name Prediction Correct - ALL Dishes) ===")
    else:
        print(f"\n=== {title} Analysis ===")
        
    print(f"Total samples: {count}")
    print(f"Min: {min_val:.2f}%")
    print(f"Max: {max_val:.2f}%")
    print(f"Mean: {mean_val:.2f}%")
    
    if is_ingredient:
        if data_list:
            high_conf = sum(1 for item in data_list if item.get('confidence') == "✅ High")
            contained_conf = sum(1 for item in data_list if item.get('confidence') == "🟡 Contained")
            print(f"Confidence distribution:")
            print(f"  - ✅ High: {high_conf} ({high_conf/count*100:.1f}%)")
            print(f"  - 🟡 Contained: {contained_conf} ({contained_conf/count*100:.1f}%)")
    
    # Create uniform width bins
    x_positions, bin_ranges, bin_labels = create_uniform_bins(values)
    
    if len(bin_ranges) == 0:
        print("No data to plot!")
        return min_val, max_val, mean_val, 0
    
    print(f"Number of bins: {len(bin_ranges)}")
    
    # Assign data to bins by category
    bin_counts_under, bin_counts_accurate, bin_counts_over = assign_data_to_bins_by_category(values, bin_ranges)
    
    # Calculate accurate prediction count
    accurate_count = sum(bin_counts_accurate)
    print(f"Accurate predictions (within ±0.1%): {accurate_count} ({accurate_count/count*100:.1f}%)")
    
    
    bars_under = ax.bar(x_positions, bin_counts_under, color='blue', edgecolor='black', alpha=0.8, width=0.8, label='Underestimation (<0)')
    
    
    bottom_for_accurate = np.array(bin_counts_under)
    bars_accurate = ax.bar(x_positions, bin_counts_accurate, bottom=bottom_for_accurate, 
                          color='green', edgecolor='black', alpha=0.8, width=0.8, label='Accurate (±0.1%)')
    
    bottom_for_over = bottom_for_accurate + np.array(bin_counts_accurate)
    bars_over = ax.bar(x_positions, bin_counts_over, bottom=bottom_for_over, 
                      color='red', edgecolor='black', alpha=0.8, width=0.8, label='Overestimation (>0)')
    
   
    ax.set_xticks(x_positions)
    ax.set_xticklabels(bin_labels, rotation=45, ha='right', fontsize=9)
    
 
    for i, (under, accurate, over) in enumerate(zip(bin_counts_under, bin_counts_accurate, bin_counts_over)):
        total = under + accurate + over
        if total > 0:
            ax.text(x_positions[i], total + 0.1, f'{total}', 
                   ha='center', va='bottom', fontsize=8)
    

    under_total = sum(bin_counts_under)
    accurate_total = sum(bin_counts_accurate)
    over_total = sum(bin_counts_over)
    

    extreme_neg = len([v for v in values if v < -1000])
    extreme_pos = len([v for v in values if v > 1000])
    
    stats_text = (
        f"Total: {count}\n"
        f"Under: {under_total} ({under_total/count*100:.1f}%)\n"
        f"Accurate: {accurate_total} ({accurate_total/count*100:.1f}%)\n"
        f"Over: {over_total} ({over_total/count*100:.1f}%)\n"
        f"Min: {min_val:.0f}%\n"
        f"Max: {max_val:.0f}%\n"
        f"Mean: {mean_val:.0f}%\n"
        f"Bins: {len(bin_ranges)}"
    )

    if is_ingredient and data_list:
        high_conf = sum(1 for item in data_list if item.get('confidence') == "✅ High")
        contained_conf = sum(1 for item in data_list if item.get('confidence') == "🟡 Contained")
        stats_text += f"\n✅ High: {high_conf}\n🟡 Contained: {contained_conf}"
    

    if extreme_neg > 0 or extreme_pos > 0:
        stats_text += f"\nExtreme (<-1000%): {extreme_neg}"
        stats_text += f"\nExtreme (>1000%): {extreme_pos}"
    
    ax.text(
        0.98, 0.98, stats_text,
        transform=ax.transAxes,
        verticalalignment='top',
        horizontalalignment='right',
        bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.8),
        fontsize=9
    )
    

    ax.legend(loc='upper left', fontsize=9, frameon=True)
    
    ax.grid(alpha=0.3, axis='y')
    ax.set_title(title, fontsize=14, fontweight='bold')
    ax.set_xlabel(xlabel, fontsize=12)
    ax.set_ylabel("Count", fontsize=12)
    

    max_count = max([sum(counts) for counts in zip(bin_counts_under, bin_counts_accurate, bin_counts_over)])
    ax.set_ylim(bottom=0, top=max_count * 1.1)
    
    return min_val, max_val, mean_val, accurate_total

# === Plot two separate histograms ===
# First plot: Per Dish Total Percentage Difference (only in correct dish)
if dish_pct_diffs:
    fig1, ax1 = plt.subplots(figsize=(18, 8))
    
    # Extract values
    values_dish = np.array([item['value'] for item in dish_pct_diffs])
    
    # 计算统计量
    min_val_dish, max_val_dish, min_dish_id, max_dish_id = find_min_max_with_ids(dish_pct_diffs)
    
    # Plot stacked histogram (green in the middle)
    min_val_dish, max_val_dish, mean_val_dish, accurate_count_dish = plot_stacked_histogram_green_in_middle(
        ax1, values_dish, 
        "Per Dish: Total Mass Percentage Difference", 
        "Percentage Difference Range",
        dish_pct_diffs,
        is_ingredient=False
    )
    
    # Add min/max information on the plot
    min_max_text = f"Min: {min_val_dish:.0f}% (dish: {min_dish_id})\nMax: {max_val_dish:.0f}% (dish: {max_dish_id})\nAccurate: {accurate_count_dish} ({accurate_count_dish/len(values_dish)*100:.1f}%)"
    ax1.text(
        0.02, 0.02, min_max_text,
        transform=ax1.transAxes,
        verticalalignment='bottom',
        bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.8),
        fontsize=9
    )
    
    # Save the first plot
    plt.tight_layout()
    dish_hist_path = input_path.replace(".xlsx", "_dish_green_middle.png")
    plt.savefig(dish_hist_path, dpi=300, bbox_inches='tight')
    print(f"\nDish histogram saved to: {dish_hist_path}")
    plt.show()

# Second plot: Per Ingredient Percentage Difference
if ingredient_pct_diffs:
    fig2, ax2 = plt.subplots(figsize=(18, 8))
    
    # Extract values
    values_ing = np.array([item['value'] for item in ingredient_pct_diffs])
    
    # Calculate statistics
    min_val_ing, max_val_ing, min_dish_id_ing, max_dish_id_ing = find_min_max_with_ids(ingredient_pct_diffs)
    
    # Plot stacked histogram (green in the middle)
    min_val_ing, max_val_ing, mean_val_ing, accurate_count_ing = plot_stacked_histogram_green_in_middle(
        ax2, values_ing,
        "Per Ingredient: Mass Percentage Difference",
        "Percentage Difference Range",
        ingredient_pct_diffs,
        is_ingredient=True
    )
    
    # Add min/max information on the plot
    min_max_text = f"Min: {min_val_ing:.0f}% (dish: {min_dish_id_ing})\nMax: {max_val_ing:.0f}% (dish: {max_dish_id_ing})\nAccurate: {accurate_count_ing} ({accurate_count_ing/len(values_ing)*100:.1f}%)"
    ax2.text(
        0.02, 0.02, min_max_text,
        transform=ax2.transAxes,
        verticalalignment='bottom',
        bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.8),
        fontsize=9
    )
    
    # Save the second plot
    plt.tight_layout()
    ingredient_hist_path = input_path.replace(".xlsx", "_ingredient_green_middle_ALL.png")
    plt.savefig(ingredient_hist_path, dpi=300, bbox_inches='tight')
    print(f"Ingredient histogram saved to: {ingredient_hist_path}")
    plt.show()

# === Overall Summary ===
print(f"\n=== Overall Summary ===")
print(f"Total correct dishes analyzed: {len(correct_dish_ids)}")
print(f"Dish percentage differences: {len(dish_pct_diffs)}")
print(f"Ingredient percentage differences (name prediction correct): {len(ingredient_pct_diffs)}")

if ingredient_pct_diffs:
    unique_dishes = set(item['dish_id'] for item in ingredient_pct_diffs)
    print(f"Unique dishes in ingredient analysis: {len(unique_dishes)}")

# Print dish_id mapping table for reference
print(f"\n=== Dish ID Reference ===")
print("Dish IDs of extreme values:")
if dish_pct_diffs:
    min_val_dish, max_val_dish, min_dish_id, max_dish_id = find_min_max_with_ids(dish_pct_diffs)
    print(f"  - Minimum total_pct_diff: {min_val_dish:.0f}% (dish: {min_dish_id})")
    print(f"  - Maximum total_pct_diff: {max_val_dish:.0f}% (dish: {max_dish_id})")

if ingredient_pct_diffs:
    min_val_ing, max_val_ing, min_dish_id_ing, max_dish_id_ing = find_min_max_with_ids(ingredient_pct_diffs)
    print(f"  - Minimum ingredient_pct_diff: {min_val_ing:.0f}% (dish: {min_dish_id_ing})")
    print(f"  - Maximum ingredient_pct_diff: {max_val_ing:.0f}% (dish: {max_dish_id_ing})")


if dish_pct_diffs:
    values_dish = [item['value'] for item in dish_pct_diffs]
    accurate_dish = [v for v in values_dish if abs(v) < 0.1]
    if accurate_dish:
        print(f"\n=== Dish Accurate Predictions (within ±0.1%) ===")
        print(f"Count: {len(accurate_dish)}/{len(values_dish)} ({len(accurate_dish)/len(values_dish)*100:.1f}%)")
        
        # Show dish_id of accurate predictions
        accurate_items = [item for item in dish_pct_diffs if abs(item['value']) < 0.1]
        if accurate_items:
            print("Accurate dish_id:")
            dish_ids = set(item['dish_id'] for item in accurate_items)
            for i, dish_id in enumerate(sorted(dish_ids)[:10], 1):  # Only show the first 10
                print(f"  {i}. dish {dish_id}")
            if len(dish_ids) > 10:
                print(f"  ... and {len(dish_ids)} more dishes")

if ingredient_pct_diffs:
    values_ing = [item['value'] for item in ingredient_pct_diffs]
    accurate_ing = [v for v in values_ing if abs(v) < 0.1]
    if accurate_ing:
        print(f"\n=== Ingredient Accurate Predictions (within ±0.1%) ===")
        print(f"Count: {len(accurate_ing)}/{len(values_ing)} ({len(accurate_ing)/len(values_ing)*100:.1f}%)")
        
        # Show dish_id of accurate predictions
        accurate_items = [item for item in ingredient_pct_diffs if abs(item['value']) < 0.1]
        if accurate_items:
            print("Accurate dish_id:")
            dish_ids = set(item['dish_id'] for item in accurate_items)
            for i, dish_id in enumerate(sorted(dish_ids)[:10], 1):  
                print(f"  {i}. dish {dish_id}")
            if len(dish_ids) > 10:
                print(f"  ... and {len(dish_ids)} more dishes")

# Print extreme value statistics
if dish_pct_diffs:
    values_dish = [item['value'] for item in dish_pct_diffs]
    extreme_neg_dish = [v for v in values_dish if v < -1000]
    extreme_pos_dish = [v for v in values_dish if v > 1000]
    
    if extreme_neg_dish or extreme_pos_dish:
        print(f"\n=== Dish Extreme Values (exceeding ±1000%) ===")
        if extreme_neg_dish:
            print(f"Negative extreme values (<-1000%): {len(extreme_neg_dish)}")
            print(f"  Range: {min(extreme_neg_dish):.0f}% to {max(extreme_neg_dish):.0f}%")
            extreme_items_neg = sorted([item for item in dish_pct_diffs if item['value'] < -1000], 
                                      key=lambda x: x['value'])[:3]
            print("  The 3 smallest negative extreme values:")
            for i, item in enumerate(extreme_items_neg, 1):
                print(f"    {i}. dish {item['dish_id']}: {item['value']:.0f}%")
        
        if extreme_pos_dish:
            print(f"Positive extreme values (>1000%): {len(extreme_pos_dish)}")
            print(f"  Range: {min(extreme_pos_dish):.0f}% to {max(extreme_pos_dish):.0f}%")
            extreme_items_pos = sorted([item for item in dish_pct_diffs if item['value'] > 1000], 
                                      key=lambda x: x['value'], reverse=True)[:3]
            print("  The 3 largest positive extreme values:")
            for i, item in enumerate(extreme_items_pos, 1):
                print(f"    {i}. dish {item['dish_id']}: {item['value']:.0f}%")

if ingredient_pct_diffs:
    values_ing = [item['value'] for item in ingredient_pct_diffs]
    extreme_neg_ing = [v for v in values_ing if v < -1000]
    extreme_pos_ing = [v for v in values_ing if v > 1000]
    
    if extreme_neg_ing or extreme_pos_ing:
        print(f"\n=== Ingredient Extreme Values (exceeding ±1000%) ===")
        if extreme_neg_ing:
            print(f"Negative extreme values (<-1000%): {len(extreme_neg_ing)}")
            print(f"  Range: {min(extreme_neg_ing):.0f}% to {max(extreme_neg_ing):.0f}%")
            extreme_items_neg = sorted([item for item in ingredient_pct_diffs if item['value'] < -1000], 
                                      key=lambda x: x['value'])[:3]
            print("  The 3 smallest negative extreme values:")
            for i, item in enumerate(extreme_items_neg, 1):
                print(f"    {i}. dish {item['dish_id']}: {item['value']:.0f}%")
        
        if extreme_pos_ing:
            print(f"Positive extreme values (>1000%): {len(extreme_pos_ing)}")
            print(f"  Range: {min(extreme_pos_ing):.0f}% to {max(extreme_pos_ing):.0f}%")
            extreme_items_pos = sorted([item for item in ingredient_pct_diffs if item['value'] > 1000], 
                                      key=lambda x: x['value'], reverse=True)[:3]
            print("  The 3 largest positive extreme values:")
            for i, item in enumerate(extreme_items_pos, 1):
                print(f"    {i}. dish {item['dish_id']}: {item['value']:.0f}%")

if not dish_pct_diffs and not ingredient_pct_diffs:
    print("❌ No valid data found! Please check:")
    print("   - File path and column indices")
    print("   - Prediction correctness values (should be 'Correct')")
    print("   - Confidence values (should be '✅ High' or '🟡 Contained')")
    print("   - Available percentage difference data")
############################################################################################

import openpyxl
import matplotlib.pyplot as plt
from statistics import mean
import numpy as np

# === User Configuration ===
input_path = r"C:\Users\PC\OneDrive\Desktop\Nutrify\food-app\Evaluation\Nutrition5k_295\Nutritio5k_295.xlsx"

# === Column Definitions (1-based indexing) ===
COLUMN_DISH_ID = 1                    # Column A: dish_id
COLUMN_PREDICTION_CORRECTNESS = 2     # Column B: Prediction_correctness
COLUMN_AVG_PCT_DIFF = 4               # Column D: avg_pct_diff (dish-level)
COLUMN_PCT_DIFF = 12                  # Column M: pct_diff (ingredient-level)
COLUMN_CONFIDENCE = 13                # Column N: confidence

print("=== Loading Excel File ===")
print(f"File path: {input_path}")
print("Column definitions:")
print(f"  - dish_id (Column A): index {COLUMN_DISH_ID}")
print(f"  - Prediction_correctness (Column B): index {COLUMN_PREDICTION_CORRECTNESS}")
print(f"  - avg_pct_diff (Column D): index {COLUMN_AVG_PCT_DIFF}")
print(f"  - pct_diff (Column M): index {COLUMN_PCT_DIFF}")
print(f"  - confidence (Column N): index {COLUMN_CONFIDENCE}")
print("=" * 50)

# === Load File ===
wb = openpyxl.load_workbook(input_path, data_only=True)
ws = wb.active

print(f"Worksheet info: rows={ws.max_row}, columns={ws.max_column}")

# Print column headers for verification
print("\n=== Column Headers ===")
headers = []
for col in range(1, 14):  # Only print first 13 columns
    header = ws.cell(row=1, column=col).value
    headers.append(header)
    print(f"Column {col}: {header}")

# === Collect Data ===
dish_pct_diffs = []      # avg_pct_diff for each correct dish
ingredient_pct_diffs = [] # pct_diff for all ingredients with confidence "✅ High" or "🟡 Contained"
correct_dish_ids = set() # dish_ids where Prediction_correctness is "Correct"

print("\n=== Collecting data ===")
current_dish_id = None
dish_avg_pct_diff_collected = False  # Flag to track if avg_pct_diff has been collected for current dish

# Iterate through all rows
for row in range(2, ws.max_row + 1):
    # Read dish_id (handling merged cells)
    cell_dish_id = ws.cell(row=row, column=COLUMN_DISH_ID).value
    if cell_dish_id is not None:
        current_dish_id = cell_dish_id
        dish_avg_pct_diff_collected = False  # Reset flag
    
    # Read other data
    correctness = ws.cell(row=row, column=COLUMN_PREDICTION_CORRECTNESS).value
    avg_pct_diff = ws.cell(row=row, column=COLUMN_AVG_PCT_DIFF).value
    pct_diff = ws.cell(row=row, column=COLUMN_PCT_DIFF).value
    confidence = ws.cell(row=row, column=COLUMN_CONFIDENCE).value
    
    # 1. Collect dish-level avg_pct_diff (only once per correct dish)
    if (correctness == "Correct" and 
        not dish_avg_pct_diff_collected and 
        avg_pct_diff is not None and 
        avg_pct_diff != ""):
        try:
            dish_pct_diffs.append({
                'value': float(avg_pct_diff),
                'dish_id': current_dish_id,
                'row': row
            })
            correct_dish_ids.add(current_dish_id)
            dish_avg_pct_diff_collected = True
        except ValueError:
            pass
    
    # 2. Collect ingredient-level pct_diff (for all ingredients with confidence "✅ High" or "🟡 Contained")
    #    Regardless of dish correctness
    if (confidence in ["✅ High", "🟡 Contained"] and 
        pct_diff is not None and 
        pct_diff != ""):
        try:
            ingredient_pct_diffs.append({
                'value': float(pct_diff),
                'dish_id': current_dish_id,
                'row': row,
                'confidence': confidence
            })
        except ValueError:
            continue

print("\n=== Data Collection Results ===")
print(f"Correct dishes found: {len(correct_dish_ids)}")
print(f"Dish avg_pct_diff samples: {len(dish_pct_diffs)}")
print(f"Ingredient pct_diff samples (name prediction correct): {len(ingredient_pct_diffs)}")

# Statistics for confidence distribution
if ingredient_pct_diffs:
    high_conf = sum(1 for item in ingredient_pct_diffs if item.get('confidence') == "✅ High")
    contained_conf = sum(1 for item in ingredient_pct_diffs if item.get('confidence') == "🟡 Contained")
    print(f"  - ✅ High confidence: {high_conf} ingredients ({high_conf/len(ingredient_pct_diffs)*100:.1f}%)")
    print(f"  - 🟡 Contained confidence: {contained_conf} ingredients ({contained_conf/len(ingredient_pct_diffs)*100:.1f}%)")
    
    # Count unique dishes involved
    unique_dishes = set(item['dish_id'] for item in ingredient_pct_diffs)
    print(f"  - Unique dishes in ingredient analysis: {len(unique_dishes)}")

# Print sample data
print(f"\nSample correct dish IDs (first 10): {list(correct_dish_ids)[:10]}")
if dish_pct_diffs:
    print(f"\nSample dish avg_pct_diffs (first 5):")
    for i, item in enumerate(dish_pct_diffs[:5]):
        print(f"  {i+1}. dish {item['dish_id']}: {item['value']:.2f}%")
if ingredient_pct_diffs:
    print(f"\nSample ingredient pct_diffs (first 5, name correct):")
    for i, item in enumerate(ingredient_pct_diffs[:5]):
        print(f"  {i+1}. dish {item['dish_id']}: {item['value']:.2f}% (confidence: {item['confidence']})")

print("\n" + "=" * 50)

# === Helper Functions ===
def find_min_max_with_ids(data_list):
    """Find minimum and maximum values with corresponding dish_ids from data list"""
    if not data_list:
        return None, None, None, None
    
    min_val = min(item['value'] for item in data_list)
    max_val = max(item['value'] for item in data_list)
    
    # Find all dish_ids corresponding to min and max values
    min_dish_ids = [item['dish_id'] for item in data_list if abs(item['value'] - min_val) < 1e-6]
    max_dish_ids = [item['dish_id'] for item in data_list if abs(item['value'] - max_val) < 1e-6]
    
    min_dish_id_str = min_dish_ids[0] if min_dish_ids else "N/A"
    max_dish_id_str = max_dish_ids[0] if max_dish_ids else "N/A"
    
    return min_val, max_val, min_dish_id_str, max_dish_id_str

def create_uniform_bins(data_values):
    """
    Create uniform width bins:
    1. Center region: [-200%, 200%] with 10% bin width
    2. Outer regions: 100% bin width
    3. Values beyond ±1000% merged into one bin
    All bins displayed with equal width in chart
    """
    if len(data_values) == 0:
        return [], [], []
    
    # Extract values
    values = np.array(data_values)
    min_val = min(values)
    max_val = max(values)
    
    # Collect all bin ranges and labels
    bin_ranges = []  # Actual range for each bin (start, end)
    bin_labels = []   # Display label for each bin
    x_positions = []  # X-axis position for each bin (starting from 0)
    
    # 1. Handle extreme negative values (< -1000%)
    if min_val < -1000:
        # Merge all values < -1000% into one bin
        extreme_neg_start = np.floor(min_val / 100) * 100  # Round down to nearest 100%
        extreme_neg_end = -1000
        bin_ranges.append((extreme_neg_start, extreme_neg_end))
        bin_labels.append(f"< -1000%\n[{int(extreme_neg_start)}%, -1000%]")
        x_positions.append(len(bin_ranges) - 1)
    
    # 2. Handle negative outer region (-1000% to -200%)
    if min_val < -200:
        neg_outer_start = max(-1000, np.floor(min_val / 100) * 100)
        neg_outer_end = -200
        
        # Create bins with 100% width
        current = neg_outer_start
        while current < neg_outer_end:
            bin_end = min(current + 100, neg_outer_end)
            bin_ranges.append((current, bin_end))
            bin_labels.append(f"[{int(current)}%, {int(bin_end)}%]")
            x_positions.append(len(bin_ranges) - 1)
            current = bin_end
    
    # 3. Center region (-200% to 200%)
    # Ensure -200% and 200% are included
    center_start = -200
    center_end = 200
    center_bin_width = 10
    
    # Create bins with 10% width
    current = center_start
    while current < center_end:
        bin_end = min(current + center_bin_width, center_end)
        bin_ranges.append((current, bin_end))
        
        # Simplify labels: show detailed label every 50%, otherwise just start value
        if current % 50 == 0 or current == -200 or current == 200 - center_bin_width:
            bin_labels.append(f"[{int(current)}%, {int(bin_end)}%]")
        else:
            bin_labels.append(f"{int(current)}%")
        
        x_positions.append(len(bin_ranges) - 1)
        current = bin_end
    
    # 4. Handle positive outer region (200% to 1000%)
    if max_val > 200:
        pos_outer_start = 200
        pos_outer_end = min(1000, np.ceil(max_val / 100) * 100)
        
        # Create bins with 100% width
        current = pos_outer_start
        while current < pos_outer_end:
            bin_end = min(current + 100, pos_outer_end)
            bin_ranges.append((current, bin_end))
            bin_labels.append(f"[{int(current)}%, {int(bin_end)}%]")
            x_positions.append(len(bin_ranges) - 1)
            current = bin_end
    
    # 5. Handle extreme positive values (> 1000%)
    if max_val > 1000:
        # Merge all values > 1000% into one bin
        extreme_pos_start = 1000
        extreme_pos_end = np.ceil(max_val / 100) * 100  # Round up to nearest 100%
        bin_ranges.append((extreme_pos_start, extreme_pos_end))
        bin_labels.append(f"> 1000%\n[1000%, {int(extreme_pos_end)}%]")
        x_positions.append(len(bin_ranges) - 1)
    
    return x_positions, bin_ranges, bin_labels

def assign_data_to_bins_by_category(values, bin_ranges):
    """
    Assign data to bins by category, return three category counts per bin
    Categories: Underestimation(<0), Accurate(0), Overestimation(>0)
    """
    bin_counts_under = [0] * len(bin_ranges)   # Underestimation
    bin_counts_accurate = [0] * len(bin_ranges) # Accurate
    bin_counts_over = [0] * len(bin_ranges)    # Overestimation
    
    for val in values:
        # Determine category
        if abs(val) < 0.1:  # Accurate prediction (within ±0.1%)
            category = 'accurate'
        elif val < 0:       # Underestimation
            category = 'under'
        else:               # Overestimation
            category = 'over'
        
        # Assign to corresponding bin
        for i, (bin_start, bin_end) in enumerate(bin_ranges):
            if bin_start <= val < bin_end or (i == len(bin_ranges) - 1 and val == bin_end):
                if category == 'under':
                    bin_counts_under[i] += 1
                elif category == 'accurate':
                    bin_counts_accurate[i] += 1
                else:  # 'over'
                    bin_counts_over[i] += 1
                break
    
    return bin_counts_under, bin_counts_accurate, bin_counts_over

def plot_stacked_histogram_green_in_middle(ax, data_values, title, xlabel, data_list=None, is_ingredient=False):
    """
    Plot stacked histogram with green in the middle
    Stacking order: Blue (underestimation) at bottom, Green (accurate) in middle, Red (overestimation) on top
    """
    if len(data_values) == 0:
        return None, None, None, None
    
    values = np.array(data_values)
    
    # Calculate statistics
    min_val = min(values)
    max_val = max(values)
    mean_val = round(mean(values), 3)
    count = len(values)
    
    if is_ingredient:
        print(f"\n=== Ingredient-Level Analysis (Name Prediction Correct - ALL Dishes) ===")
    else:
        print(f"\n=== {title} Analysis ===")
        
    print(f"Total samples: {count}")
    print(f"Minimum: {min_val:.2f}%")
    print(f"Maximum: {max_val:.2f}%")
    print(f"Mean: {mean_val:.2f}%")
    
    if is_ingredient:
        # For ingredient analysis, show confidence distribution
        if data_list:
            high_conf = sum(1 for item in data_list if item.get('confidence') == "✅ High")
            contained_conf = sum(1 for item in data_list if item.get('confidence') == "🟡 Contained")
            print(f"Confidence distribution:")
            print(f"  - ✅ High: {high_conf} ({high_conf/count*100:.1f}%)")
            print(f"  - 🟡 Contained: {contained_conf} ({contained_conf/count*100:.1f}%)")
    
    # Create uniform width bins
    x_positions, bin_ranges, bin_labels = create_uniform_bins(values)
    
    if len(bin_ranges) == 0:
        print("No data to plot!")
        return min_val, max_val, mean_val, 0
    
    print(f"Number of bins: {len(bin_ranges)}")
    
    # Assign data to bins by category
    bin_counts_under, bin_counts_accurate, bin_counts_over = assign_data_to_bins_by_category(values, bin_ranges)
    
    # Count accurate predictions
    accurate_count = sum(bin_counts_accurate)
    print(f"Accurate predictions (within ±0.1%): {accurate_count} ({accurate_count/count*100:.1f}%)")
    
    # Plot stacked bars - green in the middle
    # First plot underestimation (blue, bottom)
    bars_under = ax.bar(x_positions, bin_counts_under, color='blue', edgecolor='black', alpha=0.8, width=0.8, label='Underestimation (<0)')
    
    # Then plot accurate (green, middle) - stacked on underestimation
    bottom_for_accurate = np.array(bin_counts_under)
    bars_accurate = ax.bar(x_positions, bin_counts_accurate, bottom=bottom_for_accurate, 
                          color='green', edgecolor='black', alpha=0.8, width=0.8, label='Accurate (±0.1%)')
    
    # Finally plot overestimation (red, top) - stacked on accurate
    bottom_for_over = bottom_for_accurate + np.array(bin_counts_accurate)
    bars_over = ax.bar(x_positions, bin_counts_over, bottom=bottom_for_over, 
                      color='red', edgecolor='black', alpha=0.8, width=0.8, label='Overestimation (>0)')
    
    # Set x-axis ticks
    ax.set_xticks(x_positions)
    ax.set_xticklabels(bin_labels, rotation=45, ha='right', fontsize=9)
    
    # Add total count label for each bar
    for i, (under, accurate, over) in enumerate(zip(bin_counts_under, bin_counts_accurate, bin_counts_over)):
        total = under + accurate + over
        if total > 0:
            ax.text(x_positions[i], total + 0.1, f'{total}', 
                   ha='center', va='bottom', fontsize=8)
    
    # Statistics box
    under_total = sum(bin_counts_under)
    accurate_total = sum(bin_counts_accurate)
    over_total = sum(bin_counts_over)
    
    # Count extreme values
    extreme_neg = len([v for v in values if v < -1000])
    extreme_pos = len([v for v in values if v > 1000])
    
    stats_text = (
        f"Total: {count}\n"
        f"Under: {under_total} ({under_total/count*100:.1f}%)\n"
        f"Accurate: {accurate_total} ({accurate_total/count*100:.1f}%)\n"
        f"Over: {over_total} ({over_total/count*100:.1f}%)\n"
        f"Min: {min_val:.0f}%\n"
        f"Max: {max_val:.0f}%\n"
        f"Mean: {mean_val:.0f}%\n"
        f"Bins: {len(bin_ranges)}"
    )
    
    # For ingredient analysis, add confidence information
    if is_ingredient and data_list:
        high_conf = sum(1 for item in data_list if item.get('confidence') == "✅ High")
        contained_conf = sum(1 for item in data_list if item.get('confidence') == "🟡 Contained")
        stats_text += f"\n✅ High: {high_conf}\n🟡 Contained: {contained_conf}"
    
    # Add extreme value information if present
    if extreme_neg > 0 or extreme_pos > 0:
        stats_text += f"\nExtreme (<-1000%): {extreme_neg}"
        stats_text += f"\nExtreme (>1000%): {extreme_pos}"
    
    ax.text(
        0.98, 0.98, stats_text,
        transform=ax.transAxes,
        verticalalignment='top',
        horizontalalignment='right',
        bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.8),
        fontsize=9
    )
    
    # Add legend
    ax.legend(loc='upper left', fontsize=9, frameon=True)
    
    ax.grid(alpha=0.3, axis='y')
    ax.set_title(title, fontsize=14, fontweight='bold')
    ax.set_xlabel(xlabel, fontsize=12)
    ax.set_ylabel("Count", fontsize=12)
    
    # Set y-axis starting from 0 with some space for labels
    max_count = max([sum(counts) for counts in zip(bin_counts_under, bin_counts_accurate, bin_counts_over)])
    ax.set_ylim(bottom=0, top=max_count * 1.1)
    
    return min_val, max_val, mean_val, accurate_total

# === Plot Two Separate Histograms ===
# First plot: Per Dish Average Percentage Difference (only in correct dishes)
if dish_pct_diffs:
    fig1, ax1 = plt.subplots(figsize=(18, 8))
    
    # Extract values
    values_dish = np.array([item['value'] for item in dish_pct_diffs])
    
    # Calculate statistics
    min_val_dish, max_val_dish, min_dish_id, max_dish_id = find_min_max_with_ids(dish_pct_diffs)
    
    # Plot stacked histogram (green in the middle)
    min_val_dish, max_val_dish, mean_val_dish, accurate_count_dish = plot_stacked_histogram_green_in_middle(
        ax1, values_dish, 
        "Per Dish: Average Mass Percentage Difference", 
        "Average Percentage Difference Range",
        dish_pct_diffs,
        is_ingredient=False
    )
    
    # Add min/max information to chart
    min_max_text = f"Min: {min_val_dish:.0f}% (dish: {min_dish_id})\nMax: {max_val_dish:.0f}% (dish: {max_dish_id})\nAccurate: {accurate_count_dish} ({accurate_count_dish/len(values_dish)*100:.1f}%)"
    ax1.text(
        0.02, 0.02, min_max_text,
        transform=ax1.transAxes,
        verticalalignment='bottom',
        bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.8),
        fontsize=9
    )
    
    # Save first plot
    plt.tight_layout()
    dish_hist_path = input_path.replace(".xlsx", "_dish_avg_green_middle.png")
    plt.savefig(dish_hist_path, dpi=300, bbox_inches='tight')
    print(f"\nDish avg histogram saved to: {dish_hist_path}")
    plt.show()

# Second plot: Per Ingredient Percentage Difference (all dishes)
if ingredient_pct_diffs:
    fig2, ax2 = plt.subplots(figsize=(18, 8))
    
    # Extract values
    values_ing = np.array([item['value'] for item in ingredient_pct_diffs])
    
    # Calculate statistics
    min_val_ing, max_val_ing, min_dish_id_ing, max_dish_id_ing = find_min_max_with_ids(ingredient_pct_diffs)
    
    # Plot stacked histogram (green in the middle)
    min_val_ing, max_val_ing, mean_val_ing, accurate_count_ing = plot_stacked_histogram_green_in_middle(
        ax2, values_ing,
        "Per Ingredient: Mass Percentage Difference",
        "Percentage Difference Range",
        ingredient_pct_diffs,
        is_ingredient=True
    )
    
    # Add min/max information to chart
    min_max_text = f"Min: {min_val_ing:.0f}% (dish: {min_dish_id_ing})\nMax: {max_val_ing:.0f}% (dish: {max_dish_id_ing})\nAccurate: {accurate_count_ing} ({accurate_count_ing/len(values_ing)*100:.1f}%)"
    ax2.text(
        0.02, 0.02, min_max_text,
        transform=ax2.transAxes,
        verticalalignment='bottom',
        bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.8),
        fontsize=9
    )
    
    # Save second plot
    plt.tight_layout()
    ingredient_hist_path = input_path.replace(".xlsx", "_ingredient_green_middle_ALL.png")
    plt.savefig(ingredient_hist_path, dpi=300, bbox_inches='tight')
    print(f"Ingredient histogram saved to: {ingredient_hist_path}")
    plt.show()

# === Detailed Statistical Analysis ===
print("\n" + "=" * 50)
print("DETAILED STATISTICAL ANALYSIS")
print("=" * 50)

if dish_pct_diffs:
    values_dish = [item['value'] for item in dish_pct_diffs]
    
    print("\n--- Dish-Level Analysis (Correct Dishes Only) ---")
    print(f"Total correct dishes analyzed: {len(values_dish)}")
    print(f"Minimum average percentage difference: {min(values_dish):.2f}%")
    print(f"Maximum average percentage difference: {max(values_dish):.2f}%")
    print(f"Mean average percentage difference: {np.mean(values_dish):.2f}%")
    print(f"Median average percentage difference: {np.median(values_dish):.2f}%")
    print(f"Standard deviation: {np.std(values_dish):.2f}%")
    
    # Accuracy analysis
    for threshold in [10, 20, 50]:
        accurate = len([v for v in values_dish if abs(v) <= threshold])
        print(f"  Within ±{threshold}%: {accurate} dishes ({accurate/len(values_dish)*100:.1f}%)")
    
    # Direction analysis
    underestimations = len([v for v in values_dish if v < 0])
    overestimations = len([v for v in values_dish if v > 0])
    exact = len([v for v in values_dish if v == 0])
    
    print(f"\nDirection analysis:")
    print(f"  Underestimations (negative): {underestimations} ({underestimations/len(values_dish)*100:.1f}%)")
    print(f"  Overestimations (positive): {overestimations} ({overestimations/len(values_dish)*100:.1f}%)")
    print(f"  Exact predictions (0%): {exact} ({exact/len(values_dish)*100:.1f}%)")

if ingredient_pct_diffs:
    values_ing = [item['value'] for item in ingredient_pct_diffs]
    
    print("\n--- Ingredient-Level Analysis (Name Prediction Correct - ALL Dishes) ---")
    print(f"Total ingredients analyzed: {len(values_ing)}")
    print(f"Minimum percentage difference: {min(values_ing):.2f}%")
    print(f"Maximum percentage difference: {max(values_ing):.2f}%")
    print(f"Mean percentage difference: {np.mean(values_ing):.2f}%")
    print(f"Median percentage difference: {np.median(values_ing):.2f}%")
    print(f"Standard deviation: {np.std(values_ing):.2f}%")
    
    # Accuracy analysis
    for threshold in [10, 20, 50]:
        accurate = len([v for v in values_ing if abs(v) <= threshold])
        print(f"  Within ±{threshold}%: {accurate} ingredients ({accurate/len(values_ing)*100:.1f}%)")
    
    underestimations = len([v for v in values_ing if v < 0])
    overestimations = len([v for v in values_ing if v > 0])
    exact = len([v for v in values_ing if v == 0])
    
    print(f"\nDirection analysis:")
    print(f"  Underestimations (negative): {underestimations} ({underestimations/len(values_ing)*100:.1f}%)")
    print(f"  Overestimations (positive): {overestimations} ({overestimations/len(values_ing)*100:.1f}%)")
    print(f"  Exact predictions (0%): {exact} ({exact/len(values_ing)*100:.1f}%)")
    
    # Confidence distribution
    print(f"\nConfidence distribution (ingredient name prediction):")
    confidence_counts = {}
    for item in ingredient_pct_diffs:
        conf = item.get('confidence', 'Unknown')
        confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
    
    for conf, count in confidence_counts.items():
        print(f"  {conf}: {count} ingredients ({count/len(ingredient_pct_diffs)*100:.1f}%)")
    
    # Statistics by dish_id
    print(f"\n--- Dish Coverage for Ingredient Analysis ---")
    dish_coverage = set(item['dish_id'] for item in ingredient_pct_diffs)
    print(f"Total unique dishes included in ingredient analysis: {len(dish_coverage)}")
    print(f"Sample dish IDs (first 10): {list(dish_coverage)[:10]}")

# Summary
print("\n" + "=" * 50)
print("SUMMARY")
print("=" * 50)
print(f"Dish-level analysis (avg_pct_diff):")
print(f"  - Dishes with 'Correct' prediction: {len(correct_dish_ids)}")
print(f"  - Dish-level avg_pct_diff analyzed: {len(dish_pct_diffs)}")

print(f"\nIngredient-level analysis (pct_diff):")
print(f"  - Ingredient-level pct_diff analyzed: {len(ingredient_pct_diffs)}")
print(f"  - Analysis includes ALL dishes (not restricted to correct dishes)")
print(f"  - Only ingredients with name prediction confidence: '✅ High' or '🟡 Contained'")

if not dish_pct_diffs and not ingredient_pct_diffs:
    print("\n❌ No valid data found! Please check:")
    print("   1. File path is correct")
    print("   2. Excel file has the expected columns")
    print("   3. 'Prediction_correctness' column has 'Correct' values (for dish analysis)")
    print("   4. 'confidence' column has '✅ High' or '🟡 Contained' values (for ingredient analysis)")

    print("   5. Percentage difference columns have numeric values")
